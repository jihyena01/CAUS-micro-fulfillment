from neural_embedded_model import createLRP, write_to_txt_cvrplib_format
from solver_cvrp import dataCvrp
import os
import argparse
import openpyxl
from dataparse import create_data
from datetime import datetime
import logging
import sys
from openpyxl import Workbook, load_workbook

parser = argparse.ArgumentParser(description='Process arguments for LRP script.')
parser.add_argument('--BFS', type=str, help='Path to save integer feasible solutions directory')
parser.add_argument('--phi_loc', type=str, help='Path to phi model file')
parser.add_argument('--rho_loc', type=str, help='Path to rho model file')
parser.add_argument('--existing_excel_file', type=str, help='Path to the existing Excel file')
parser.add_argument('--sheet_name', type=str, help='Name of the sheet in the Excel file')
parser.add_argument('--normalization', type=str, choices=['fixed', 'dynamic'], default='dynamic', help='Normalization type: choose "fixed" for using a fixed fi value, or "dynamic" for calculating fi dynamically.')

args = parser.parse_args()
BFS = args.BFS
phi_loc = args.phi_loc
rho_loc = args.rho_loc
existing_excel_file = args.existing_excel_file
sheet_name = args.sheet_name
fi_mode_input =  args.normalization

log_dir = "log_files/mip_nn"
os.makedirs(log_dir, exist_ok=True)
# Directory containing the prodhon dataset
directory_path = "NEO-LRP/prodhon_dataset"

try:
    workbook = load_workbook(existing_excel_file)
except FileNotFoundError:
    print(f"Excel file not found. Creating new file: {existing_excel_file}")
    workbook = Workbook()
    workbook.save(existing_excel_file)

if sheet_name not in workbook.sheetnames:
    workbook.create_sheet(sheet_name)
worksheet = workbook[sheet_name]

headings = [
    "Instance", "FLP", "VRP", "LRP(MIP+NN)", "NumRoutes_OptSol",
    "Exec time per depot(MIP+NN)", "initial solution generation time",
    "NN model execution time", "VRPSolverEasy computed VRP cost",
    "actual LRP cost(using VRPSolverEasy)",
    "avg solver_cvrp script execution time per depot",
    "total solver_cvrp script execution time",
    "VRPSolverEasy model solve time", "BKS",
    "Optimization_gap_optsol", "Prediction_gap"
]

if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet.cell(1, 1).value is None:
    # Add headings to the first row
    for col, heading in enumerate(headings, start=1):
        worksheet.cell(row=1, column=col, value=heading)

def has_customers(instance_file_path):
    """Check if a VRP instance has any customers assigned."""
    with open(instance_file_path, 'r') as file:
        lines = file.readlines()
    try:
        demand_section_index = lines.index("DEMAND_SECTION\n")
        depot_section_index = lines.index("DEPOT_SECTION\n")
    except ValueError as e:
        raise ValueError(f"Section missing in file {instance_file_path}: {e}")
    demand_lines = lines[demand_section_index + 1:depot_section_index]
    # Remove depot demand (assumed to be the first line)
    customer_demands = demand_lines[1:]
    return len(customer_demands) > 0

bks_dict = {
    "coord20-5-1.dat": 54793,
    "coord20-5-1b.dat": 39104,
    "coord20-5-2.dat": 48908,
    "coord20-5-2b.dat": 37542,
    "coord50-5-1.dat": 90111,
    "coord50-5-1b.dat": 63242,
    "coord50-5-2.dat": 88293,
    "coord50-5-2b.dat": 67308,
    "coord50-5-2bBIS.dat": 51822,
    "coord50-5-2BIS.dat": 84055,
    "coord50-5-3.dat": 86203,
    "coord50-5-3b.dat": 61830,
    "coord100-5-1.dat": 274814,
    "coord100-5-1b.dat": 213568,
    "coord100-5-2.dat": 193671,
    "coord100-5-2b.dat": 157095,
    "coord100-5-3.dat": 200079,
    "coord100-5-3b.dat": 152441,
    "coord100-10-1.dat": 287661,
    "coord100-10-1b.dat": 230989,
    "coord100-10-2.dat": 243590,
    "coord100-10-2b.dat": 203988,
    "coord100-10-3.dat": 250882,
    "coord100-10-3b.dat": 203114,
    "coord200-10-1.dat": 474850,
    "coord200-10-1b.dat": 375177,
    "coord200-10-2.dat": 448077,
    "coord200-10-2b.dat": 373696,
    "coord200-10-3.dat": 469433,
    "coord200-10-3b.dat": 362320
}

# the list of instances already in the excel
processed_instances = set()
for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
    instance_name = row[0]
    if instance_name is not None:
        processed_instances.add(instance_name)
        
for filename in bks_dict.keys(): 
    if filename in processed_instances:
        print(f"Instance {filename} already processed. Skipping.")
        continue

    # Initialize lists to collect metrics for averaging
    flp_cost_list = []
    vrp_cost_list = []
    lrp_cost_list = []
    vrp_routes_optsol_list = []
    lrp_exec_list = []
    warmstart_time_list = []
    nn_model_time_list = []
    vrp_easy_vrp_cost_list = []
    actual_lrp_cost_list = []
    ve_exec_list = []
    tot_ve_exec_list = []
    vrp_solver_easy_model_solve_time_list = []
    gap_list = []
    gap_vrp_perc_list = []
        
    file_path = os.path.join(directory_path, filename)
    if os.path.exists(file_path):
        print("Working on:", file_path)
    else:
        print("File not found:", file_path)
        break

    for run_index in range(5):
        print(f"Run {run_index + 1} for instance {filename}")

        # Define per-run subdirectory
        instance_subdir_run = os.path.join(BFS, os.path.splitext(filename)[0], f"run_{run_index}")
        os.makedirs(instance_subdir_run, exist_ok=True)

        # Prepare logging, etc.
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        log_filename = f"{os.path.splitext(filename)[0]}_{current_time}_run_{run_index}.log"

        print(f'\n\n Working on file :{file_path}')
        ans = create_data(file_path)  # Process the data

        # Timing and solver execution
        lrp_st = datetime.now()
        lrp_solver = createLRP(ans)
        lrp_result = lrp_solver.model(
            file_path,
            log_filename,
            instance_subdir_run,
            phi_loc,
            rho_loc,
            fi_mode=fi_mode_input,
            fixed_fi_value=1000.0
        )
        lrp_ed = datetime.now()

        warmstart_time = lrp_result[4]
        nn_model_time = lrp_result[5]

        # Process assignment results
        flp_dict = {}
        for j in range(len(lrp_result[0])):
            if lrp_result[0][j] > 0.5:
                ls = []
                for i in range(len(lrp_result[1][j])):
                    if lrp_result[1][j][i] > 0.5:
                        ls.append(i)
                flp_dict[j] = ls

        rout_dist = {}
        fac_cust_dem = {}
        cust_dem_fac = {}
        for f in flp_dict:
            ls1 = []
            ls2 = []
            dem_sum = 0
            for c in flp_dict[f]:
                ls1.append(ans[3][c])
                dem_sum += ans[6][c]
                ls2.append(ans[6][c])
            ls1.insert(0, ans[2][f])
            ls2.insert(0, 0)
            rout_dist[f] = ls1
            fac_cust_dem[f] = dem_sum
            cust_dem_fac[f] = ls2
        ass_result = [lrp_result[2], flp_dict, rout_dist, fac_cust_dem, cust_dem_fac]

        ve_st = datetime.now()
        vrpeasy_solver = dataCvrp(ans, ass_result)
        vrp_easy_results = vrpeasy_solver.runVRPeasy()
        ve_ed = datetime.now()

        od = len(flp_dict)

        lrp_exec = ((lrp_ed - lrp_st).total_seconds()) / od
        warmstart_time = warmstart_time / od
        tot_ve_exec = (ve_ed - ve_st).total_seconds()
        ve_exec = tot_ve_exec / od

        instance_name = os.path.basename(file_path)
        bks = bks_dict.get(instance_name)

        flp_cost = lrp_result[2]
        vrp_cost = lrp_result[3]
        lrp_cost = flp_cost + vrp_cost
        actual_lrp_cost = vrp_easy_results[0]
        vrp_easy_vrp_cost = vrp_easy_results[1]
        vrp_routes_optsol = sum(vrp_easy_results[3])

        if bks is not None and bks != 0:
            gap = (abs(bks - actual_lrp_cost) / bks) * 100
            gap = round(gap, 2)
        else:
            gap = "N/A"

        if lrp_result[3] != 0 and vrp_easy_vrp_cost != 0:
            gap_vrp_perc = ((vrp_easy_vrp_cost - lrp_result[3]) / vrp_easy_vrp_cost) * 100
            gap_vrp_perc = abs(round(gap_vrp_perc, 2))
        else:
            gap_vrp_perc = "N/A"

        vrp_solver_easy_model_solve_time = vrp_easy_results[7]

        # Append metrics to lists
        flp_cost_list.append(flp_cost)
        vrp_cost_list.append(vrp_cost)
        lrp_cost_list.append(lrp_cost)
        vrp_routes_optsol_list.append(vrp_routes_optsol)
        lrp_exec_list.append(lrp_exec)
        warmstart_time_list.append(warmstart_time)
        nn_model_time_list.append(nn_model_time)
        vrp_easy_vrp_cost_list.append(vrp_easy_vrp_cost)
        actual_lrp_cost_list.append(actual_lrp_cost)
        ve_exec_list.append(ve_exec)
        tot_ve_exec_list.append(tot_ve_exec)
        vrp_solver_easy_model_solve_time_list.append(vrp_solver_easy_model_solve_time)
        gap_list.append(gap if gap != "N/A" else 0)
        gap_vrp_perc_list.append(gap_vrp_perc if gap_vrp_perc != "N/A" else 0)

    # Compute averages
    avg_flp_cost = sum(flp_cost_list) / len(flp_cost_list)
    avg_vrp_cost = sum(vrp_cost_list) / len(vrp_cost_list)
    avg_lrp_cost = sum(lrp_cost_list) / len(lrp_cost_list)
    avg_vrp_routes_optsol = sum(vrp_routes_optsol_list) / len(vrp_routes_optsol_list)
    avg_lrp_exec = sum(lrp_exec_list) / len(lrp_exec_list)
    avg_warmstart_time = sum(warmstart_time_list) / len(warmstart_time_list)
    avg_nn_model_time = sum(nn_model_time_list) / len(nn_model_time_list)
    avg_vrp_easy_vrp_cost = sum(vrp_easy_vrp_cost_list) / len(vrp_easy_vrp_cost_list)
    avg_actual_lrp_cost = sum(actual_lrp_cost_list) / len(actual_lrp_cost_list)
    avg_ve_exec = sum(ve_exec_list) / len(ve_exec_list)
    avg_tot_ve_exec = sum(tot_ve_exec_list) / len(tot_ve_exec_list)
    avg_vrp_solver_easy_model_solve_time = sum(vrp_solver_easy_model_solve_time_list) / len(vrp_solver_easy_model_solve_time_list)
    avg_gap = sum(gap_list) / len(gap_list) if gap_list else "N/A"
    avg_gap_vrp_perc = sum(gap_vrp_perc_list) / len(gap_vrp_perc_list) if gap_vrp_perc_list else "N/A"

    # Build new_row
    new_row = [
        os.path.basename(file_path),
        avg_flp_cost,
        avg_vrp_cost,
        avg_lrp_cost,
        avg_vrp_routes_optsol,
        avg_lrp_exec,
        avg_warmstart_time,
        avg_nn_model_time,
        avg_vrp_easy_vrp_cost,
        avg_actual_lrp_cost,
        avg_ve_exec,
        avg_tot_ve_exec,
        avg_vrp_solver_easy_model_solve_time,
        bks,
        avg_gap,
        avg_gap_vrp_perc
    ]

    worksheet.append(new_row)
    workbook.save(existing_excel_file)